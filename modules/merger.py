import io
import re
import pandas as pd
import numpy as np
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ── Column classification ──────────────────────────────────────────────────────
EARNING_COLS  = ["BASIC", "HOUSING", "TRANSPORT", "ALLOWANCE", "BONUS",
                 "OVERTIME", "COMMISSION", "GRATUITY", "13TH MONTH",
                 "OTHER EARNINGS", "ARREARS"]
DEDUCTION_COLS = ["TAX", "PENSION", "LOAN", "SAL. ADV.", "SALARY ADVANCE",
                  "PENALTY", "NHIS", "NHF", "COOPERATIVE", "UNION DUES",
                  "OTHER DEDUCTIONS", "TOTAL DED.", "DEDUCTION"]
SKIP_COLS      = ["STAFF NAME", "_KEY", "_GROSS", "NET SALARY", "TOTAL DED.",
                  "GROSS SALARY", "GROSS"]


def _normalize_name(name: str) -> str:
    """Lowercase, strip, collapse internal spaces."""
    return re.sub(r"\s+", " ", str(name).strip().lower())


def load_single_file(file, filename: str) -> tuple[pd.DataFrame, list]:
    """
    Load one Excel file. Returns (clean_df, validation_issues).
    clean_df has a _KEY column (normalised name) and _SOURCE column.
    """
    issues = []
    try:
        df = pd.read_excel(file, engine="openpyxl")
    except Exception as exc:
        return pd.DataFrame(), [f"[{filename}] Cannot read file: {exc}"]

    # Clean column names
    df.columns = [str(c).strip().upper() for c in df.columns]

    # Drop fully blank rows / columns
    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    if "STAFF NAME" not in df.columns:
        issues.append(f"[{filename}] Missing STAFF NAME column — file skipped.")
        return pd.DataFrame(), issues

    # Validate names
    for idx, val in df["STAFF NAME"].items():
        raw = str(val).strip()
        if not raw or raw.lower() in ("nan", "none", ""):
            issues.append(f"[{filename}] Row {idx+2}: blank STAFF NAME — row skipped.")
        elif df["STAFF NAME"].str.lower().str.strip().value_counts().get(raw.lower(), 0) > 1:
            issues.append(f"[{filename}] Possible duplicate name: '{raw}'")

    # Drop blank name rows
    df["STAFF NAME"] = df["STAFF NAME"].astype(str).str.strip()
    df = df[df["STAFF NAME"].str.lower().isin(
        [n for n in df["STAFF NAME"].str.lower() if n not in ("nan", "none", "")]
    )].copy()

    # Normalise all numeric columns
    for col in df.columns:
        if col != "STAFF NAME":
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Validate numeric values
    for col in df.columns:
        if col != "STAFF NAME":
            neg = df[df[col] < 0]
            if not neg.empty:
                issues.append(f"[{filename}] Column '{col}' has {len(neg)} negative value(s).")

    df["_KEY"]    = df["STAFF NAME"].apply(_normalize_name)
    df["_SOURCE"] = filename

    return df, issues


def merge_files(
    file_list: list[tuple],        # [(file_obj, filename), ...]
    dup_strategy: str = "sum",     # "sum" | "first" | "last"
    progress_cb=None,              # callable(pct, msg)
) -> dict:
    """
    Load and merge multiple payroll files.

    Returns dict with keys:
      merged_df, audit_df, validation_issues, file_summaries,
      earning_cols, deduction_cols
    """
    all_dfs      = []
    audit_rows   = []
    val_issues   = []
    file_summaries = []
    n_files      = len(file_list)

    # ── Step 1: Load all files ─────────────────────────────────────────────
    for i, (fobj, fname) in enumerate(file_list):
        if progress_cb:
            progress_cb(int((i / n_files) * 40), f"📂 Loading {fname}…")

        df, issues = load_single_file(fobj, fname)
        val_issues.extend(issues)

        if df.empty:
            continue

        file_summaries.append({
            "File": fname,
            "Employees": len(df),
            "Columns": [c for c in df.columns if c not in ("STAFF NAME", "_KEY", "_SOURCE")],
        })

        # Build audit rows
        for _, row in df.iterrows():
            for col in df.columns:
                if col in ("STAFF NAME", "_KEY", "_SOURCE"):
                    continue
                audit_rows.append({
                    "Source File":       fname,
                    "Employee Name":     row["STAFF NAME"],
                    "Payroll Component": col,
                    "Imported Amount":   row[col],
                })

        all_dfs.append(df)

    if not all_dfs:
        return {
            "merged_df": pd.DataFrame(),
            "audit_df": pd.DataFrame(),
            "validation_issues": val_issues,
            "file_summaries": file_summaries,
            "earning_cols": [],
            "deduction_cols": [],
        }

    if progress_cb:
        progress_cb(45, "🔗 Merging employee records…")

    # ── Step 2: Merge all DataFrames ───────────────────────────────────────
    combined = pd.concat(all_dfs, ignore_index=True, sort=False)
    combined.fillna(0, inplace=True)

    # Collect all payload columns (exclude meta)
    meta = {"STAFF NAME", "_KEY", "_SOURCE"}
    payload_cols = [c for c in combined.columns if c not in meta]

    if progress_cb:
        progress_cb(60, "➕ Applying duplicate column strategy…")

    # ── Step 3: Group by normalised key ───────────────────────────────────
    def _agg(group: pd.DataFrame) -> pd.Series:
        result = {}
        # Canonical name = first non-blank occurrence (title-cased)
        result["STAFF NAME"] = (
            group["STAFF NAME"].iloc[0].strip().title()
        )
        for col in payload_cols:
            if col not in group.columns:
                result[col] = 0.0
                continue
            vals = pd.to_numeric(group[col], errors="coerce").fillna(0)
            if dup_strategy == "sum":
                result[col] = float(vals.sum())
            elif dup_strategy == "first":
                nz = vals[vals != 0]
                result[col] = float(nz.iloc[0]) if not nz.empty else 0.0
            else:  # last
                nz = vals[vals != 0]
                result[col] = float(nz.iloc[-1]) if not nz.empty else 0.0
        return pd.Series(result)

    merged = combined.groupby("_KEY", sort=False).apply(_agg).reset_index(drop=True)

    if progress_cb:
        progress_cb(75, "🧮 Calculating gross, deductions, net…")

    # ── Step 4: Classify columns ──────────────────────────────────────────
    data_cols = [c for c in merged.columns if c != "STAFF NAME"]

    earn_cols = []
    ded_cols  = []
    for c in data_cols:
        cu = c.upper()
        if any(e in cu for e in [
            "BASIC","HOUSING","TRANSPORT","ALLOWANCE","BONUS",
            "OVERTIME","COMMISSION","GRATUITY","13TH","ARREARS","EARNING"
        ]):
            earn_cols.append(c)
        elif any(d in cu for d in [
            "TAX","PENSION","LOAN","ADV","PENALTY","NHIS","NHF",
            "COOPERATIVE","UNION","DEDUCTION","DED."
        ]):
            ded_cols.append(c)
        # columns that don't match either stay in data but aren't double-counted

    # Remove legacy computed cols if present
    for drop in ["TOTAL DED.", "NET SALARY", "GROSS SALARY", "_GROSS", "GROSS"]:
        if drop in merged.columns:
            merged.drop(columns=[drop], inplace=True)
        earn_cols = [c for c in earn_cols if c != drop]
        ded_cols  = [c for c in ded_cols  if c != drop]

    # Recompute totals
    merged["GROSS SALARY"]      = merged[earn_cols].sum(axis=1) if earn_cols else 0
    merged["TOTAL DED."]        = merged[ded_cols].sum(axis=1)  if ded_cols  else 0
    merged["NET SALARY"]        = merged["GROSS SALARY"] - merged["TOTAL DED."]

    if progress_cb:
        progress_cb(90, "📋 Building audit log…")

    audit_df = pd.DataFrame(audit_rows) if audit_rows else pd.DataFrame(
        columns=["Source File","Employee Name","Payroll Component","Imported Amount"]
    )

    if progress_cb:
        progress_cb(100, "✅ Merge complete!")

    return {
        "merged_df":        merged,
        "audit_df":         audit_df,
        "validation_issues": val_issues,
        "file_summaries":   file_summaries,
        "earning_cols":     earn_cols,
        "deduction_cols":   ded_cols,
    }


# ── Export helpers ─────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Consolidated Payroll")
        ws = writer.sheets["Consolidated Payroll"]
        from openpyxl.styles import Font, PatternFill, Alignment
        orange = PatternFill("solid", fgColor="E8610A")
        wfont  = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill   = orange
            cell.font   = wfont
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)
    buf.seek(0)
    return buf.read()


def export_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def export_report_pdf(df: pd.DataFrame, settings: dict) -> bytes:
    """Generate a landscape PDF master payroll report."""
    import base64
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    ORANGE    = colors.HexColor("#E8610A")
    DARK_GRAY = colors.HexColor("#2C2C2C")
    LIGHT_GRAY= colors.HexColor("#F5F5F5")
    MID_GRAY  = colors.HexColor("#CCCCCC")
    WHITE     = colors.white
    ALT_ROW   = colors.HexColor("#FFF3EC")

    s_title = ParagraphStyle("t", fontSize=14, fontName="Helvetica-Bold",
                             textColor=DARK_GRAY, alignment=TA_LEFT)
    s_sub   = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",
                             textColor=DARK_GRAY, alignment=TA_LEFT)
    s_foot  = ParagraphStyle("f", fontSize=7,  fontName="Helvetica",
                             textColor=MID_GRAY,  alignment=TA_CENTER)

    story = []

    # Header row
    logo_b64 = settings.get("logo_b64")
    hdr_cells = []
    if logo_b64:
        from reportlab.platypus import Image as RLImage
        img_buf = io.BytesIO(base64.b64decode(logo_b64))
        hdr_cells.append(RLImage(img_buf, width=30*mm, height=10*mm, kind="proportional"))
    else:
        hdr_cells.append(Paragraph("", s_title))

    hdr_cells.append(
        Paragraph(
            f"{settings.get('company_name','')}<br/>"
            f"<font size=9>{settings.get('company_address','')}</font>",
            s_title,
        )
    )
    hdr_cells.append(
        Paragraph(
            f"CONSOLIDATED PAYROLL REPORT<br/>"
            f"<font size=9>{settings.get('payroll_month','')} {settings.get('payroll_year','')}</font>",
            ParagraphStyle("tr", fontSize=11, fontName="Helvetica-Bold",
                           textColor=ORANGE, alignment=TA_LEFT),
        )
    )
    hdr_t = Table([hdr_cells], colWidths=["15%","45%","40%"])
    hdr_t.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",(0,0),(-1,-1),4),
    ]))
    story.append(hdr_t)
    story.append(HRFlowable(width="100%", thickness=2, color=ORANGE, spaceAfter=4*mm))

    # Build table data
    # Only show key columns to fit landscape A4
    key_cols = ["STAFF NAME","GROSS SALARY","TOTAL DED.","NET SALARY"]
    extra = [c for c in df.columns if c not in key_cols]
    show_cols = key_cols[:1] + extra[:6] + key_cols[1:]  # name | extras | totals
    show_cols = [c for c in show_cols if c in df.columns]

    s_hdr  = ParagraphStyle("h", fontSize=7.5, fontName="Helvetica-Bold",
                            textColor=WHITE, alignment=TA_CENTER)
    s_cell = ParagraphStyle("c", fontSize=7.5, fontName="Helvetica",
                            textColor=DARK_GRAY, alignment=TA_LEFT)
    s_num  = ParagraphStyle("n", fontSize=7.5, fontName="Helvetica",
                            textColor=DARK_GRAY, alignment=TA_RIGHT)

    def _fmt(val, col):
        if col == "STAFF NAME":
            return Paragraph(str(val), s_cell)
        try:
            return Paragraph(f"₦{float(val):,.0f}", s_num)
        except:
            return Paragraph(str(val), s_cell)

    tbl_data = [[Paragraph(c.replace("."," "), s_hdr) for c in show_cols]]
    for _, row in df.iterrows():
        tbl_data.append([_fmt(row.get(c, 0), c) for c in show_cols])

    page_w = landscape(A4)[0] - 24*mm
    name_w = 45*mm
    other_w = (page_w - name_w) / (len(show_cols) - 1)
    col_widths = [name_w] + [other_w] * (len(show_cols) - 1)

    tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    ts  = [
        ("BACKGROUND",(0,0),(-1,0), ORANGE),
        ("TEXTCOLOR",(0,0),(-1,0), WHITE),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica"),
        ("FONTSIZE",(0,0),(-1,-1),7.5),
        ("TOPPADDING",(0,0),(-1,-1),4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),5),
        ("RIGHTPADDING",(0,0),(-1,-1),5),
        ("LINEBELOW",(0,0),(-1,-1),0.4, MID_GRAY),
        ("BACKGROUND",(0,-1),(-1,-1), LIGHT_GRAY),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
    ]
    for i in range(1, len(tbl_data)-1, 2):
        ts.append(("BACKGROUND",(0,i),(-1,i), ALT_ROW))
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    story.append(Spacer(1,4*mm))

    # Totals summary
    tot_earn = df["GROSS SALARY"].sum() if "GROSS SALARY" in df.columns else 0
    tot_ded  = df["TOTAL DED."].sum()   if "TOTAL DED."   in df.columns else 0
    tot_net  = df["NET SALARY"].sum()   if "NET SALARY"   in df.columns else 0
    summary_data = [
        [Paragraph("TOTAL PAYROLL SUMMARY", ParagraphStyle("ts",fontSize=9,
         fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_LEFT)), "", "", ""],
        ["Total Employees", f"{len(df)}", "Total Gross", f"₦{tot_earn:,.2f}"],
        ["Total Deductions", f"₦{tot_ded:,.2f}", "Total Net Salary", f"₦{tot_net:,.2f}"],
    ]
    sum_t = Table(summary_data, colWidths=["25%","25%","25%","25%"])
    sum_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), ORANGE),
        ("SPAN",(0,0),(-1,0)),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("BACKGROUND",(0,1),(-1,-1), LIGHT_GRAY),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica-Bold"),
        ("LINEBELOW",(0,0),(-1,-1),0.4, MID_GRAY),
    ]))
    story.append(sum_t)

    from datetime import datetime as dt
    story.append(Spacer(1,3*mm))
    story.append(HRFlowable(width="100%",thickness=1,color=MID_GRAY))
    story.append(Paragraph(
        f"Generated by Smart Payroll Payslip Generator | {dt.now().strftime('%d %B %Y %H:%M')} | Confidential",
        s_foot
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()
